

from io import BytesIO

import pandas as pd
import streamlit as st

try:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


LLM_API_KEY = ""
LLM_MODEL = ""
LLM_BASE_URL = ""


def llm_is_configured() -> bool:
    return bool(LLM_API_KEY.strip()) and bool(LLM_MODEL.strip())


def llm_explain_changes(prompt: str) -> str:
    return ()


st.set_page_config(
    page_title="Data Comparison Agent",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

MAX_FILES = 4
DIFF_COLOR = "FFEB9C"
MISSING_COLOR = "FFC7CE"
SAME_COLOR = "C6EFCE"
HEADER_COLOR = "4472C4"
UP_COLOR = "C6EFCE"
DOWN_COLOR = "FFC7CE"


if "file_slots" not in st.session_state:
    st.session_state.file_slots = 2


def add_slot():
    if st.session_state.file_slots < MAX_FILES:
        st.session_state.file_slots += 1


def remove_slot():
    if st.session_state.file_slots > 2:
        st.session_state.file_slots -= 1


def load_file(uploaded_file):
    if uploaded_file is None:
        return None
    name = uploaded_file.name.lower()
    try:
        if name.endswith(".csv"):
            return pd.read_csv(uploaded_file)
        if name.endswith((".xlsx", ".xls")):
            return pd.read_excel(uploaded_file)
        if name.endswith(".xml"):
            return pd.read_xml(uploaded_file)
    except Exception as exc:
        st.error(f"'{uploaded_file.name}' dosyası okunamadı: {exc}")
    return None


def ensure_unique(name, existing):
    if name not in existing:
        return name
    n = 2
    while f"{name} ({n})" in existing:
        n += 1
    return f"{name} ({n})"


def build_column_matrix(dfs_with_names):
    all_cols = []
    for _, df in dfs_with_names:
        for c in df.columns:
            if c not in all_cols:
                all_cols.append(c)
    rows = []
    for col in all_cols:
        row = {"Kolon adı": col}
        present = 0
        for name, df in dfs_with_names:
            if col in df.columns:
                row[name] = "✓ Var"
                present += 1
            else:
                row[name] = "✗ Yok"
        row["Durum"] = (
            "Tüm dosyalarda var"
            if present == len(dfs_with_names)
            else "Bazı dosyalarda yok"
        )
        rows.append(row)
    return pd.DataFrame(rows)


def build_wide_comparison(dfs_with_names, key_col):
    file_names = [name for name, _ in dfs_with_names]

    all_cols = []
    for _, df in dfs_with_names:
        for c in df.columns:
            if c != key_col and c not in all_cols:
                all_cols.append(c)

    indexed = {}
    for name, df in dfs_with_names:
        d = df.copy()
        d[key_col] = d[key_col].astype(str)
        d = d.drop_duplicates(subset=[key_col]).set_index(key_col)
        indexed[name] = d

    all_keys = set()
    for d in indexed.values():
        all_keys.update(d.index.tolist())
    all_keys = sorted(all_keys)

    value_rows = []
    status_rows = []

    for k in all_keys:
        v_row = {"Satır anahtarı": k}
        s_row = {"Satır anahtarı": ""}
        row_has_diff = False
        row_has_missing = False

        for col in all_cols:
            cell_values = []
            for name in file_names:
                d = indexed[name]
                if col in d.columns and k in d.index:
                    v = d.at[k, col]
                    if pd.isna(v):
                        cell_values.append(None)
                    else:
                        cell_values.append(v)
                else:
                    cell_values.append(None)

            non_null_str = [str(v) for v in cell_values if v is not None]
            col_is_diff = len(set(non_null_str)) > 1 if non_null_str else False
            col_has_missing = any(v is None for v in cell_values)

            if col_is_diff:
                row_has_diff = True
            if col_has_missing:
                row_has_missing = True

            for i, name in enumerate(file_names):
                label = f"{col} — {name}"
                v_row[label] = "" if cell_values[i] is None else cell_values[i]
                if cell_values[i] is None:
                    s_row[label] = "missing"
                elif col_is_diff:
                    s_row[label] = "diff"
                else:
                    s_row[label] = "same"

        if row_has_diff:
            v_row["Durum"] = "Farklılık var"
            s_row["Durum"] = "diff"
        elif row_has_missing:
            v_row["Durum"] = "Bazı dosyalarda yok"
            s_row["Durum"] = "missing"
        else:
            v_row["Durum"] = "Tamamen aynı"
            s_row["Durum"] = "same"

        value_rows.append(v_row)
        status_rows.append(s_row)

    values_df = pd.DataFrame(value_rows)
    status_df = pd.DataFrame(status_rows)

    summary = {
        "total_keys": len(all_keys),
        "identical": int((values_df["Durum"] == "Tamamen aynı").sum()) if not values_df.empty else 0,
        "with_diff": int((values_df["Durum"] == "Farklılık var").sum()) if not values_df.empty else 0,
        "missing": int((values_df["Durum"] == "Bazı dosyalarda yok").sum()) if not values_df.empty else 0,
    }

    return values_df, status_df, summary


def build_metric_comparison(dfs_with_names, key_col, metric_col, measure="pct"):
    """Her ID için ilk dosyadaki değer ile diğer dosyalardaki değeri
    doğrudan karşılaştırır (toplam/ortalama yok). Aynı ID birden fazla
    satırda geçiyorsa ilk satır alınır ve uyarı üretilir."""
    file_names = [name for name, _ in dfs_with_names]

    per_file_series = {}
    duplicate_warnings = {}
    for name, df in dfs_with_names:
        if key_col not in df.columns or metric_col not in df.columns:
            per_file_series[name] = pd.Series(dtype=float)
            continue
        d = df[[key_col, metric_col]].copy()
        d[key_col] = d[key_col].astype(str)
        d[metric_col] = pd.to_numeric(d[metric_col], errors="coerce")
        dup_count = int(d.duplicated(subset=[key_col]).sum())
        if dup_count > 0:
            duplicate_warnings[name] = dup_count
        d = d.drop_duplicates(subset=[key_col], keep="first")
        per_file_series[name] = d.set_index(key_col)[metric_col]

    all_keys = sorted(set().union(*[s.index for s in per_file_series.values()]))

    rows = []
    for k in all_keys:
        row = {key_col: k}
        for name in file_names:
            s = per_file_series[name]
            if k in s.index and pd.notna(s[k]):
                row[name] = float(s[k])
            else:
                row[name] = None
        rows.append(row)
    per_key_df = pd.DataFrame(rows)

    baseline = file_names[0]
    for name in file_names[1:]:
        diff_col = f"Fark ({name} − {baseline})"
        pct_col = f"% değişim ({name} vs {baseline})"
        ratio_col = f"Oran ({name} / {baseline})"

        def _diff(r, n=name):
            a, b = r[baseline], r[n]
            if a is None or b is None:
                return None
            return b - a

        def _pct(r, n=name):
            a, b = r[baseline], r[n]
            if a is None or b is None or a == 0:
                return None
            return (b - a) / a * 100.0

        def _ratio(r, n=name):
            a, b = r[baseline], r[n]
            if a is None or b is None or a == 0:
                return None
            return b / a

        per_key_df[diff_col] = per_key_df.apply(_diff, axis=1)
        per_key_df[pct_col] = per_key_df.apply(_pct, axis=1)
        per_key_df[ratio_col] = per_key_df.apply(_ratio, axis=1)

    totals = {
        name: float(pd.Series(per_file_series[name]).dropna().sum())
        for name in file_names
    }

    summary = {
        "key_col": key_col,
        "metric_col": metric_col,
        "measure": measure,
        "baseline": baseline,
        "file_names": file_names,
        "totals": totals,
        "num_keys": len(all_keys),
        "duplicate_warnings": duplicate_warnings,
    }
    return per_key_df, summary


def static_change_analysis(per_key_df, summary, top_n=5):
    file_names = summary["file_names"]
    baseline = summary["baseline"]
    metric = summary["metric_col"]
    key_col = summary["key_col"]
    totals = summary["totals"]

    lines = []
    lines.append(
        f"**Baz dosya:** `{baseline}` — karşılaştırma değeri: `{metric}` ({key_col} bazlı)"
    )
    for name in file_names[1:]:
        pct_col = f"% değişim ({name} vs {baseline})"
        if pct_col not in per_key_df.columns:
            continue
        sub = per_key_df[[key_col, baseline, name, pct_col]].dropna(subset=[pct_col])
        if sub.empty:
            lines.append(f"- `{name}`: karşılaştırılabilir ID bulunamadı.")
            continue
        n_up = int((sub[pct_col] > 0).sum())
        n_down = int((sub[pct_col] < 0).sum())
        n_flat = int((sub[pct_col] == 0).sum())
        avg_pct = float(sub[pct_col].mean())
        yön = "ortalamada arttı 📈" if avg_pct > 0 else ("ortalamada azaldı 📉" if avg_pct < 0 else "ortalamada değişmedi ➖")
        lines.append(
            f"- `{name}`: {len(sub)} ID karşılaştırıldı — {n_up} artış, {n_down} azalış, "
            f"{n_flat} değişmedi. Ortalama % değişim: {avg_pct:+.2f}% — {yön}"
        )

    lines.append("")
    lines.append("**ID bazlı öne çıkanlar:**")
    for name in file_names[1:]:
        pct_col = f"% değişim ({name} vs {baseline})"
        if pct_col not in per_key_df.columns:
            continue
        sub = per_key_df[[key_col, baseline, name, pct_col]].dropna(subset=[pct_col])
        if sub.empty:
            continue
        top_up = sub.nlargest(top_n, pct_col)
        top_down = sub.nsmallest(top_n, pct_col)
        lines.append(f"\n*{name}* — en çok artan {min(top_n, len(top_up))}:")
        for _, r in top_up.iterrows():
            lines.append(
                f"  - {key_col}={r[key_col]}: {r[baseline]:,.2f} → {r[name]:,.2f} "
                f"({r[pct_col]:+.2f}%)"
            )
        lines.append(f"\n*{name}* — en çok azalan {min(top_n, len(top_down))}:")
        for _, r in top_down.iterrows():
            lines.append(
                f"  - {key_col}={r[key_col]}: {r[baseline]:,.2f} → {r[name]:,.2f} "
                f"({r[pct_col]:+.2f}%)"
            )

    return "\n".join(lines)


def style_wide(values_df, status_df):
    v = values_df.reset_index(drop=True)
    s = status_df.reset_index(drop=True)

    def apply_styles(df):
        styles = pd.DataFrame("", index=df.index, columns=df.columns)
        for col in df.columns:
            if col not in s.columns:
                continue
            for ri in df.index:
                status = s.at[ri, col]
                if status == "diff":
                    styles.at[ri, col] = f"background-color: #{DIFF_COLOR}"
                elif status == "missing":
                    styles.at[ri, col] = f"background-color: #{MISSING_COLOR}"
                elif status == "same" and col == "Durum":
                    styles.at[ri, col] = f"background-color: #{SAME_COLOR}"
        return styles

    return v.style.apply(apply_styles, axis=None)


def style_metric_table(per_key_df):
    def color_pct(val):
        if pd.isna(val):
            return ""
        if val > 0:
            return f"background-color: #{UP_COLOR}"
        if val < 0:
            return f"background-color: #{DOWN_COLOR}"
        return ""

    signed_cols = [
        c for c in per_key_df.columns
        if c.startswith("% değişim") or c.startswith("Fark (")
    ]
    styler = per_key_df.style
    if signed_cols:
        styler = styler.map(color_pct, subset=signed_cols)
    fmt = {}
    for c in per_key_df.columns:
        if per_key_df[c].dtype.kind in "fi":
            fmt[c] = "{:,.2f}"
    if fmt:
        styler = styler.format(fmt, na_rep="—")
    return styler


def style_column_matrix(col_matrix_df):
    def apply_styles(df):
        styles = pd.DataFrame("", index=df.index, columns=df.columns)
        for ri in df.index:
            durum = df.at[ri, "Durum"]
            if durum == "Bazı dosyalarda yok":
                styles.at[ri, "Durum"] = f"background-color: #{MISSING_COLOR}"
            elif durum == "Tüm dosyalarda var":
                styles.at[ri, "Durum"] = f"background-color: #{SAME_COLOR}"
            for col in df.columns:
                if col in ("Kolon adı", "Durum"):
                    continue
                if df.at[ri, col] == "✗ Yok":
                    styles.at[ri, col] = f"background-color: #{MISSING_COLOR}"
        return styles

    return col_matrix_df.style.apply(apply_styles, axis=None)


def _style_workbook(wb, values_df, status_df, col_matrix_df):
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    diff_fill = PatternFill("solid", fgColor=DIFF_COLOR)
    missing_fill = PatternFill("solid", fgColor=MISSING_COLOR)
    same_fill = PatternFill("solid", fgColor=SAME_COLOR)
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 14
            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, 45))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len
        ws.row_dimensions[1].height = 28

    def paint_from_status(ws, value_df, status_subset):
        col_headers = [c.value for c in ws[1]]
        for row_idx in range(len(value_df)):
            for c_idx, col_name in enumerate(col_headers, start=1):
                if col_name not in status_subset.columns:
                    continue
                status = status_subset.iloc[row_idx][col_name]
                cell = ws.cell(row=row_idx + 2, column=c_idx)
                if col_name == "Durum":
                    if status == "diff":
                        cell.fill = diff_fill
                    elif status == "missing":
                        cell.fill = missing_fill
                    elif status == "same":
                        cell.fill = same_fill
                else:
                    if status == "diff":
                        cell.fill = diff_fill
                    elif status == "missing":
                        cell.fill = missing_fill

    if "Tum Satirlar" in wb.sheetnames:
        paint_from_status(wb["Tum Satirlar"], values_df, status_df)
        wb["Tum Satirlar"].freeze_panes = "B2"

    if "Sadece Farkliliklar" in wb.sheetnames:
        diff_mask = values_df["Durum"] != "Tamamen aynı"
        diff_values = values_df[diff_mask].reset_index(drop=True)
        diff_status = status_df[diff_mask].reset_index(drop=True)
        paint_from_status(wb["Sadece Farkliliklar"], diff_values, diff_status)
        wb["Sadece Farkliliklar"].freeze_panes = "B2"

    if "Kolon Karsilastirmasi" in wb.sheetnames:
        ws = wb["Kolon Karsilastirmasi"]
        headers = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2):
            for idx, col_name in enumerate(headers):
                cell = row[idx]
                if col_name == "Durum":
                    if cell.value == "Bazı dosyalarda yok":
                        cell.fill = missing_fill
                    elif cell.value == "Tüm dosyalarda var":
                        cell.fill = same_fill
                elif col_name not in ("Kolon adı",):
                    if cell.value == "✗ Yok":
                        cell.fill = missing_fill


def build_excel(dfs_with_names, col_matrix_df, values_df, status_df, summary):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        ozet = pd.DataFrame(
            [
                ["Yüklenen dosya sayısı", len(dfs_with_names)],
                ["Toplam benzersiz satır sayısı", summary["total_keys"]],
                ["Tüm dosyalarda aynı olan satırlar", summary["identical"]],
                ["Farklılık içeren satırlar", summary["with_diff"]],
                ["Bazı dosyalarda olmayan satırlar", summary["missing"]],
                ["", ""],
                ["Renk açıklaması", ""],
                ["Sarı", "Aynı satırın değeri dosyalar arasında farklı"],
                ["Kırmızı", "Değer bazı dosyalarda yok"],
                ["Yeşil", "Tüm dosyalarda aynı"],
            ],
            columns=["Açıklama", "Değer"],
        )
        ozet.to_excel(writer, sheet_name="Ozet", index=False)

        col_matrix_df.to_excel(writer, sheet_name="Kolon Karsilastirmasi", index=False)
        values_df.to_excel(writer, sheet_name="Tum Satirlar", index=False)

        only_diffs = values_df[values_df["Durum"] != "Tamamen aynı"].reset_index(drop=True)
        only_diffs.to_excel(writer, sheet_name="Sadece Farkliliklar", index=False)

        if HAS_OPENPYXL:
            _style_workbook(writer.book, values_df, status_df, col_matrix_df)

    output.seek(0)
    return output.getvalue()


def build_metric_excel(per_key_df, summary, narrative):
    output = BytesIO()
    measure_label = {
        "diff": "Mutlak fark (B − baz)",
        "pct": "Yüzde değişim (%)",
        "ratio": "Oran (kaç kat)",
    }.get(summary.get("measure", "pct"), summary.get("measure", ""))
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        info = pd.DataFrame(
            [
                ["Eşleştirme kolonu", summary["key_col"]],
                ["Karşılaştırma değeri", summary["metric_col"]],
                ["Nasıl karşılaştırıldı", measure_label],
                ["Baz dosya", summary["baseline"]],
                ["ID sayısı", summary["num_keys"]],
            ],
            columns=["Açıklama", "Değer"],
        )
        info.to_excel(writer, sheet_name="Ozet", index=False)
        per_key_df.to_excel(writer, sheet_name="ID Bazli", index=False)
        pd.DataFrame({"Analiz": narrative.splitlines()}).to_excel(
            writer, sheet_name="Analiz", index=False
        )
    output.seek(0)
    return output.getvalue()


st.title("Data Comparison Agent")

st.markdown(
    """
    Bu araç, yüklediğiniz **2 ila 4 dosya** arasındaki farkları size net bir şekilde gösterir.
    Dosyalarınız üzerinde **hiçbir değişiklik yapılmaz**; araç sadece farklılıkları tespit eder ve görsel olarak işaretler.
    """
)

with st.expander("Nasıl kullanılır?", expanded=False):
    st.markdown(
        """
        1. Karşılaştırmak istediğiniz dosyaları yükleyin (Excel, CSV veya XML formatında).
        2. İsterseniz alttaki **➕ Karşılaştırmaya bir dosya daha ekle** butonuyla en fazla 4 dosyaya kadar ekleyebilirsiniz.
        3. **Karşılaştırma modu** seçin:
            - **Metrik karşılaştırma (varsayılan):** Ortak bir eşleştirme kolonu (örn. *şube no*) ve değişimini izlemek istediğiniz bir metrik (örn. *toplam gelir*) seçin. Yüzdesel değişim ve değişim sebepleri gösterilir.
            - **Birebir (identical) karşılaştırma:** Satırlar tüm kolon değerleriyle birebir karşılaştırılır.
        4. **Karşılaştırmayı başlat** butonuna basın.
        5. Sonuçları ekranda görebilir veya Excel dosyası olarak indirebilirsiniz.
        """
    )

st.divider()

st.header("1. Dosyalarınızı yükleyin")
st.caption("En az 2, en fazla 4 dosya yükleyebilirsiniz. Desteklenen formatlar: .csv, .xlsx, .xls, .xml")

uploaded_files = []
num_slots = st.session_state.file_slots
idx = 0
while idx < num_slots:
    row_cols = st.columns(2)
    for rc in row_cols:
        if idx >= num_slots:
            break
        with rc:
            f = st.file_uploader(
                f"Dosya {idx + 1}",
                type=["csv", "xlsx", "xls", "xml"],
                key=f"upload_{idx}",
            )
            if f is not None:
                uploaded_files.append(f)
        idx += 1

btn_cols = st.columns([2, 2, 3])
with btn_cols[0]:
    st.button(
        "➕ Karşılaştırmaya bir dosya daha ekle",
        on_click=add_slot,
        disabled=num_slots >= MAX_FILES,
        use_container_width=True,
    )
with btn_cols[1]:
    st.button(
        "➖ Son alanı kaldır",
        on_click=remove_slot,
        disabled=num_slots <= 2,
        use_container_width=True,
    )

dfs_with_names = []
for file in uploaded_files:
    df = load_file(file)
    if df is not None:
        existing_names = [n for n, _ in dfs_with_names]
        unique_name = ensure_unique(file.name, existing_names)
        dfs_with_names.append((unique_name, df))

if len(dfs_with_names) < 2:
    st.info("Karşılaştırma için en az 2 dosya yüklemeniz gerekir.")
    st.stop()

st.success(f"✅ {len(dfs_with_names)} dosya başarıyla yüklendi.")

st.divider()

st.header("2. Yüklediğiniz dosyalara göz atın")
st.caption("Aşağıdaki sekmelerden her dosyanın içeriğini kontrol edebilirsiniz.")

preview_tabs = st.tabs([name for name, _ in dfs_with_names])
for tab, (name, df) in zip(preview_tabs, dfs_with_names):
    with tab:
        mc1, mc2 = st.columns(2)
        mc1.metric("Satır sayısı", df.shape[0])
        mc2.metric("Kolon sayısı", df.shape[1])
        st.dataframe(df, use_container_width=True, hide_index=True)

st.divider()

st.header("3. Karşılaştırma modu ve kolonlar")

mode = st.radio(
    "Karşılaştırma modu",
    [
        "Metrik karşılaştırma (yüzdesel değişim)",
        "Birebir (identical) karşılaştırma",
    ],
    horizontal=False,
    help=(
        "Metrik mod: ortak bir eşleştirme kolonu (örn. şube no) üzerinden "
        "seçtiğiniz bir metriğin (örn. toplam gelir) dosyalar arası değişimini gösterir. "
        "Birebir mod: tüm kolonlardaki değerleri birebir karşılaştırır."
    ),
)

common_columns = set(dfs_with_names[0][1].columns)
for _, df in dfs_with_names[1:]:
    common_columns &= set(df.columns)
common_columns = sorted(common_columns)

if not common_columns:
    st.error(
        "❌ Yüklediğiniz dosyalarda **ortak bir kolon bulunamadı**. "
        "Satırları eşleştirebilmek için tüm dosyalarda en az bir ortak kolon olması gerekir."
    )
    col_matrix_df = build_column_matrix(dfs_with_names)
    st.dataframe(
        style_column_matrix(col_matrix_df),
        use_container_width=True,
        hide_index=True,
    )
    st.stop()

key_col = st.selectbox(
    "Eşleştirme kolonu (örn. şube no, müşteri no, sipariş no)",
    common_columns,
    help="Satırları bu kolona göre eşleştireceğiz.",
)

metric_col = None
measure = "pct"
if mode.startswith("Metrik"):
    numeric_common = []
    for c in common_columns:
        if c == key_col:
            continue
        is_num = all(
            pd.api.types.is_numeric_dtype(df[c])
            or pd.to_numeric(df[c], errors="coerce").notna().any()
            for _, df in dfs_with_names
        )
        if is_num:
            numeric_common.append(c)

    if not numeric_common:
        st.error(
            "❌ Metrik karşılaştırma için tüm dosyalarda ortak, sayısal bir kolon bulunamadı. "
            "Birebir modu deneyebilir veya dosyalarınızı kontrol edebilirsiniz."
        )
        st.stop()

    mc1, mc2 = st.columns(2)
    with mc1:
        metric_col = st.selectbox(
            "Karşılaştırma değeri (örn. toplam gelir, adet, ciro)",
            numeric_common,
            help=(
                "Her ID için ilk dosyadaki bu değer ile diğer dosyalardaki "
                "karşılığı birebir kıyaslanır."
            ),
        )
    with mc2:
        measure_label = st.selectbox(
            "Nasıl karşılaştıralım?",
            [
                "Yüzde değişim (% arttı/azaldı)",
                "Mutlak fark (ne kadar arttı)",
                "Oran (kaç kat oldu)",
            ],
            help=(
                "Sonuç tablosunda tüm ölçütler gösterilir; burada seçtiğiniz, "
                "özet ve sıralamalarda öne çıkan ölçüttür."
            ),
        )
        if measure_label.startswith("Mutlak"):
            measure = "diff"
        elif measure_label.startswith("Oran"):
            measure = "ratio"
        else:
            measure = "pct"

start = st.button(
    "🔍 Karşılaştırmayı başlat",
    type="primary",
    use_container_width=True,
)

if start:
    with st.spinner("Dosyalarınız karşılaştırılıyor, lütfen bekleyin..."):
        if mode.startswith("Metrik"):
            per_key_df, m_summary = build_metric_comparison(
                dfs_with_names, key_col, metric_col, measure=measure
            )
            narrative = static_change_analysis(per_key_df, m_summary)
            st.session_state.comparison_result = {
                "mode": "metric",
                "per_key_df": per_key_df,
                "summary": m_summary,
                "narrative": narrative,
            }
        else:
            col_matrix_df = build_column_matrix(dfs_with_names)
            values_df, status_df, summary = build_wide_comparison(dfs_with_names, key_col)
            st.session_state.comparison_result = {
                "mode": "identical",
                "col_matrix_df": col_matrix_df,
                "values_df": values_df,
                "status_df": status_df,
                "summary": summary,
            }

if "comparison_result" not in st.session_state:
    st.stop()

result = st.session_state.comparison_result

st.divider()
st.header("4. Sonuçlar")

if result["mode"] == "metric":
    per_key_df = result["per_key_df"]
    m_summary = result["summary"]
    narrative = result["narrative"]

    baseline = m_summary["baseline"]
    metric = m_summary["metric_col"]

    if m_summary.get("duplicate_warnings"):
        warn_lines = [
            f"- `{n}`: {c} tekrarlı kayıt (ilk değeri kullandık)"
            for n, c in m_summary["duplicate_warnings"].items()
        ]
        st.warning(
            "⚠️ Bazı dosyalarda aynı ID birden fazla satırda geçiyor. "
            "Karşılaştırma **ilk değer** üzerinden yapıldı:\n" + "\n".join(warn_lines)
        )

    st.subheader("Dosya toplamları (bilgi amaçlı)")
    st.caption(
        f"Baz dosya: **{baseline}** — karşılaştırma değeri: **{metric}**. "
        "Karşılaştırma ID bazlıdır; aşağıdaki toplamlar yalnızca genel büyüklük göstergesidir."
    )

    cols = st.columns(len(m_summary["file_names"]))
    base_total = m_summary["totals"][baseline]
    for c, name in zip(cols, m_summary["file_names"]):
        t = m_summary["totals"][name]
        if name == baseline:
            c.metric(name, f"{t:,.2f}")
        else:
            delta_pct = ((t - base_total) / base_total * 100.0) if base_total else 0.0
            c.metric(name, f"{t:,.2f}", f"{delta_pct:+.2f}%")

    st.subheader(f"ID bazlı ({m_summary['key_col']}) karşılaştırma")
    st.caption(
        "Her ID için **baz dosyadaki değer** ile **diğer dosyalardaki değeri** birebir kıyaslanır. "
        "Yeşil = artış, kırmızı = azalış."
    )
    st.dataframe(
        style_metric_table(per_key_df),
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("Değişim analizi")
    if llm_is_configured():
        st.caption("💡 LLM destekli yorum")
        prompt = (
            "Aşağıdaki metrik karşılaştırma sonuçlarını iş perspektifinden yorumla, "
            "olası sebepleri listele ve aksiyon öner:\n\n" + narrative
        )
        st.markdown(llm_explain_changes(prompt))
        with st.expander("Statik özet"):
            st.markdown(narrative)
    else:
        st.caption(
            "LLM yapılandırması yok — statik analiz gösteriliyor. "
            "LLM_API_KEY ve LLM_MODEL doldurulduğunda LLM yorumu otomatik aktif olur."
        )
        st.markdown(narrative)

    st.subheader("Excel indir")
    excel_bytes = build_metric_excel(per_key_df, m_summary, narrative)
    st.download_button(
        "📥 Metrik karşılaştırma Excel'i",
        data=excel_bytes,
        file_name="metrik_karsilastirma.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

else:
    col_matrix_df = result["col_matrix_df"]
    values_df = result["values_df"]
    status_df = result["status_df"]
    summary = result["summary"]

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Toplam satır", summary["total_keys"])
    m2.metric("Tamamen aynı", summary["identical"])
    m3.metric("Farklılık içeren", summary["with_diff"])
    m4.metric("Bazı dosyalarda yok", summary["missing"])

    st.markdown(
        f"""
        <div style="margin-top: 12px;">
        <span style="background-color:#{DIFF_COLOR};padding:4px 10px;border-radius:4px;margin-right:8px;">🟡 Farklılık</span>
        <span style="background-color:#{MISSING_COLOR};padding:4px 10px;border-radius:4px;margin-right:8px;">🔴 Eksik</span>
        <span style="background-color:#{SAME_COLOR};padding:4px 10px;border-radius:4px;">🟢 Aynı</span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    tab_rows, tab_cols, tab_excel = st.tabs(
        ["Tüm Satırlar (Yan Yana)", "Kolon Karşılaştırması", "Excel İndir"]
    )

    with tab_rows:
        st.markdown(
            "Aşağıda yüklediğiniz **tüm satırlar** gösteriliyor. Her satır için, "
            "aynı anahtarın **farklı dosyalardaki karşılıkları yan yana** listelenmiştir. "
            "Farklı olan değerler **sarı**, bazı dosyalarda bulunmayan değerler **kırmızı** ile işaretlenmiştir."
        )
        filter_choice = st.radio(
            "Filtrele:",
            [
                "Tümünü göster",
                "Sadece farklılık içerenler",
                "Sadece bazı dosyalarda olmayanlar",
                "Sadece tamamen aynı olanlar",
            ],
            horizontal=True,
        )
        mask = pd.Series(True, index=values_df.index)
        if filter_choice == "Sadece farklılık içerenler":
            mask = values_df["Durum"] == "Farklılık var"
        elif filter_choice == "Sadece bazı dosyalarda olmayanlar":
            mask = values_df["Durum"] == "Bazı dosyalarda yok"
        elif filter_choice == "Sadece tamamen aynı olanlar":
            mask = values_df["Durum"] == "Tamamen aynı"

        display_values = values_df[mask].reset_index(drop=True)
        display_status = status_df[mask].reset_index(drop=True)

        if display_values.empty:
            st.info("Bu filtreye uyan satır bulunamadı.")
        else:
            st.dataframe(
                style_wide(display_values, display_status),
                use_container_width=True,
                hide_index=True,
            )

    with tab_cols:
        st.markdown(
            "Her kolonun hangi dosyalarda bulunduğunu aşağıdan görebilirsiniz. "
            "**✓ Var** = kolon o dosyada mevcut, **✗ Yok** = kolon o dosyada bulunmuyor."
        )
        st.dataframe(
            style_column_matrix(col_matrix_df),
            use_container_width=True,
            hide_index=True,
        )

    with tab_excel:
        st.markdown(
            "Karşılaştırma sonucunu **renk kodlu bir Excel dosyası** olarak indirebilirsiniz. "
            "İndirdiğiniz dosyada farklılıklar **sarı**, eksik değerler **kırmızı**, "
            "aynı olanlar **yeşil** ile işaretlenmiştir."
        )
        excel_bytes = build_excel(
            dfs_with_names, col_matrix_df, values_df, status_df, summary
        )
        st.download_button(
            "📥 Excel dosyasını indir",
            data=excel_bytes,
            file_name="karsilastirma_sonucu.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
